import openpyxl

wb = openpyxl. Workbook() # 임시 엑셀 생성 
wb.create_sheet("Test") # 테스트라는 시트 추가

sheet = wb["Test"] # 시트 선택
sheet ['A1'] = 10
sheet ['B1'] = 20
sheet.cell(1, 3, 30)

sheet.cell(row=1, column=4, value=40) 
sheet.cell(1, 5,"=SUM(A1:D1)")

wb.save("/Users/parkcha/Desktop/1.Project/ForCodingTest/00.알고리즘및파이썬정리/Coding.xlsx")